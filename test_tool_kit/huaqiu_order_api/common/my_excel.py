from openpyxl import load_workbook


class MyExcel:
    def __init__(self, excel_file_path, sheet_name):
        # 1、加载一个excel，得到工作薄 Workbook
        self.wb = load_workbook(excel_file_path)
        self.sheet_name = sheet_name
        # 2、选择一个表单- 通过表单名 Sheet
        self.sh = self.wb[sheet_name]
        # self.excel_file_path = excel_file_path

    def read_data(self):
        # 接口的请求数据，读取出来是字符串
        # 存储表单下读取到的所有数据 - 每一个成员都是一个字典
        all_data = []
        data = list(self.sh.values)  # 把excel列表里面的所有值读取出来 并转换成列表
        keys = data[0]  # 获取所有的列名
        for row in data[1:]:
            row_dict = dict(zip(keys, row))  # zip(keys, row): 把keys和row拼接起来
            all_data.append(row_dict)
        return all_data

    def get_title(self):
        """
        获取接口用例中的title
        :return: title
        """
        cases = self.read_data()
        title = []
        for case in cases:
            title.append(case['title'])
        return title

    def excel_read(self, value,row,column):
        """
        Excel插入数据
        :param sheet_name_value 表头名
        ：param row 行数
        :param column 列数
        :param value 插入内容
        """

        sheet = self.wb.active
        # # 插入表头列名
        # sheet[self.sheet_name] = sheet_name_value
        # 在表格第row行第column列插入数据
        sheet.cell(row=row, column=column).value =value
        # self.wb.save(self.excel_file_path)
        return self



if __name__ == '__main__':
    # excel的文件路径
    excel_path = r"E:\HuaQiu_API\testdatas\测试用例.xlsx"
    MyExcel(excel_path, '商品加购接口').get_title()





