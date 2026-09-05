import time
from datetime import datetime

import jsonpath
import openpyxl
import requests
import yaml

from huaqiu_order_api.HQCHIP_SOO.login import SOOLogin
from huaqiu_order_api.HQCHIP_SRM.partner_audit.partner_audit import PartnerAudit
from huaqiu_order_api.HQCHIP_SRM.partner_users.partner_users import PartnerUsers
from huaqiu_order_api.HQCHIP_SRM.partner_users.users_password_reset import UsersPasswordReset
from huaqiu_order_api.HQCHIP_SRM.pass_partner.pass_partner import PassPartner
from huaqiu_order_api.common.loguru_logger import logger
from huaqiu_order_api.common.my_path import settle_goods_dir, yaml_file, account_yaml
from huaqiu_order_api.common.yaml_handler import write_yaml


class SettleGoods:
    # 库存发布
    def __init__(self, target_rss, supplier_name=None, goods_name=None, provider_name=None, encap=None, Overview=None,
                 min_pack=None, MOQ=None, min_step_price=None, inland_delivery_day=None, CNY_domestic_stock=None
                 ):
        self.srm_rss = target_rss
        self.json_head = {"Content-Type": "application/json"}
        self.file_head = {"Content-Type": "multipart/form-source_data"}
        self.supplier_name = supplier_name
        self.goods_name = goods_name
        self.provider_name = provider_name
        self.encap = encap
        self.min_pack = min_pack
        self.MOQ = MOQ
        self.Overview = Overview
        self.min_step_price_CNY = min_step_price
        self.inland_delivery_day = inland_delivery_day
        self.CNY_domestic_stock = CNY_domestic_stock
        with open(yaml_file, 'r', encoding='utf-8') as yamlfile:
            data = yaml.load(yamlfile, Loader=yaml.FullLoader)
        self.SRM_URL = data['SRM_URL']

    def excel_file_write(self):
        """更新文件里面型号、品牌、封装、最小包装数量、最小起订量、第一阶梯人民币价格、大陆交期"""
        # 打开指定的Excel文件
        logger.info("导入的供应商为: {}，导入的信息为：型号为{}，品牌为{}，封装为{}，最小包装数量为{}，最小起订量为{}，大陆库存为{}，"
                    "第一阶梯人民币价格为{}，大陆交期为{}".format(self.supplier_name, self.goods_name, self.provider_name, self.encap, self.min_pack,
                    self.MOQ, self.CNY_domestic_stock,  self.min_step_price_CNY, self.inland_delivery_day))
        workbook = openpyxl.load_workbook(settle_goods_dir)
        # 选择活动工作表
        sheet = workbook.active
        # 修改Excel文件里面的型号、品牌、封装
        sheet.cell(row=4, column=1).value = self.goods_name
        sheet.cell(row=4, column=2).value = self.provider_name
        sheet.cell(row=4, column=3).value = self.encap
        if self.Overview != None:
            sheet.cell(row=4, column=4).value = self.Overview
        if self.min_pack != None:
            sheet.cell(row=4, column=5).value = self.min_pack
        if self.MOQ != None:
            sheet.cell(row=4, column=6).value = self.MOQ
        if self.min_step_price_CNY != None:
            sheet.cell(row=4, column=7).value = self.min_step_price_CNY
            min_step_price_USD = float(self.min_step_price_CNY) / 7.1921
            sheet.cell(row=4, column=11).value = min_step_price_USD
            multiple = 1
            for i in range(1, 3):
                n = 16 + (i-1) * 3 + (i-1)
                i = i + 1
                multiple = multiple * i * 10
                logger.info(f"打印i为{i}，阶梯系数为{multiple}")
                if self.MOQ != None:
                    step_num = int(self.MOQ) * multiple
                    sheet.cell(row=4, column=n).value = step_num
                    tiered_price_CNY = float(self.min_step_price_CNY) / multiple
                    sheet.cell(row=4, column=n + 1).value = float(self.min_step_price_CNY) / multiple
                    tiered_price_USD = float(self.min_step_price_CNY) / multiple / 7.1921
                    sheet.cell(row=4, column=n + 2).value = float(self.min_step_price_CNY) / multiple / 7.1921
                    logger.info(f"生成第{i}阶梯，阶梯数量为{step_num}，阶梯价格CNY为{tiered_price_CNY}，阶梯价格USD为{tiered_price_USD}")
        if self.CNY_domestic_stock != None and self.CNY_domestic_stock != '':
            sheet.cell(row=4, column=9).value = self.CNY_domestic_stock
            sheet.cell(row=4, column=12).value = int(self.CNY_domestic_stock) / 10
        if self.inland_delivery_day != None:
            sheet.cell(row=4, column=10).value = self.inland_delivery_day
            if '-' in self.inland_delivery_day:
                min_delivery_day = self.inland_delivery_day.split('-')[0]
                max_delivery_day = self.inland_delivery_day.split('-')[1]
                hk_delivery_day = str(int(min_delivery_day) + 7) + '-' + str(int(max_delivery_day) + 7)
            else:
                hk_delivery_day = str(int(self.inland_delivery_day) + 7)
            sheet.cell(row=4, column=13).value = hk_delivery_day
        # 批号
        current_date = datetime.now()
        year = current_date.year
        sheet.cell(row=4, column=14).value = year

        # 保存修改后的Excel文件
        workbook.save(settle_goods_dir)
        return self
    def settle_search_excel(self):
        """库存发布"""

        #提交导入
        excel_goods_url = "{}/partnermanage/partnerSaleGoodsDetail/analysisExcel".format(self.SRM_URL)
        excel_goods_body ={"goodType": 1, "multipartFile": self.fileInfo ,"supplienCode":"","supplierName": self.supplier_name,"undercanriage":1}
        excel_goods_res = self.srm_rss.post(url=excel_goods_url, json=excel_goods_body,headers=self.json_head).json()
        logger.info(excel_goods_res)
        msg = ''
        if 'msg' in excel_goods_res:
            msg = excel_goods_res["msg"]
        if msg != "供应商不支持代售现货类型合作":
            self.bill_sn =excel_goods_res["body"]["billSn"]
            logger.info(f"导入单号:{self.bill_sn}")

            bill_list_url = "{}/partnermanage/partnerSaleBill/noPartnerSaleBillPage".format(self.SRM_URL)
            bill_list_body = {"billSn": self.bill_sn, "current":1, "size":10}
            bill_list_res = self.srm_rss.post(url=bill_list_url, json=bill_list_body, headers=self.json_head).json()

            # 受理
            bill_confirmOk_url = "{}/partnermanage/partnerSaleBill/confirmOk".format(self.SRM_URL)
            bill_confirmOk_body = {"billSn": self.bill_sn, "acceptanceOpinions": "test"}
            bill_confirmOk_res = self.srm_rss.post(url=bill_confirmOk_url, json=bill_confirmOk_body, headers=self.json_head).json()
            logger.info(bill_confirmOk_res)

            # 提交审核
            start_audit_url = "{}/partnermanage/startProcessSale".format(self.SRM_URL)
            start_audit_body = {"comment": "启动流程", "procDefKey": "replace_sale_goods", "subject": "【代售-库存更新审核】","supplierCode": self.bill_sn}

            start_audit_res = self.srm_rss.post(url=start_audit_url, json=start_audit_body, headers=self.json_head).json()
            logger.info(start_audit_res)
            self.settle_goods_procInstId = start_audit_res["body"]
            print(f"库存更新审核id：{self.settle_goods_procInstId}")
        return self
    def settle_goods_auditor_acquire(self, procInstId):
        # # 获取审核人
        # procInstId  库存更新审核id
        approvalrecord_url = "{}/partnermanage/approvalRecord".format(self.SRM_URL)
        approvalrecord_body = {"procInstId": procInstId}
        approvalrecord_res = self.srm_rss.post(url=approvalrecord_url, json=approvalrecord_body, headers=self.json_head).json()
        logger.info(approvalrecord_res)
        if "msg" in approvalrecord_res:
            if "token无效" in approvalrecord_res["msg"]:
                target_rss = SOOLogin("uat-srm.huaqiu.com", "partnermanage").target_login()
                self.srm_rss = target_rss
                approvalrecord_res = self.srm_rss.post(url=approvalrecord_url, json=approvalrecord_body, headers=self.json_head).json()
                logger.info(approvalrecord_res)
        approvalInfo = approvalrecord_res["body"]["list"]
        self.uidCreator = None
        if approvalInfo != []:
            creatorDesc = jsonpath.jsonpath(approvalrecord_res, "$..creatorDesc")
            uidCreator = jsonpath.jsonpath(approvalrecord_res, "$..uidCreator")
            applyStatus = jsonpath.jsonpath(approvalrecord_res, "$..applyStatus")
            for i in range(len(creatorDesc)):
                if creatorDesc[i] != "发起人":
                    if applyStatus[i] != "已通过":
                        self.uidCreator = uidCreator[i]
        logger.info(f"审核主题名称为{self.bill_sn}的审核节点为{self.uidCreator}")
        logger.info("请转至审核人的审核中心进行操作审核中心")
        return self.uidCreator

    def settle_goods_file(self):
        """上传库存"""
        settle_goods_file_url = "{}/partnermanage/partnerFile/upload".format(self.SRM_URL)
        file = [('file', ("渠道员-现货发布模版.xlsx", open(settle_goods_dir, 'rb'),'multipart/form-source_data.openxmlformats-officedocument.spreadsheetml.sheet'))]
        settle_goods_file_res = self.srm_rss.post(url=settle_goods_file_url, files=file).json()
        logger.info(settle_goods_file_res)
        self.fileInfo = settle_goods_file_res["body"]
        return self


    def parther_audit_list(self,comment="通过",operation="pass"):
        """审核中心-我的审批"""
        suc = None
        if self.bill_sn != '':
            search_url = "{}/partnermanage/queryTodo".format(self.SRM_URL)
            search_body = {"subject": self.bill_sn, "procDefKey": ["supplier_apply_into", "supplier_update_data","supplier_ban_change","supplier_taka_data","replace_sale_goods"]}
            search_res = self.srm_rss.post(url=search_url, json=search_body, headers=self.json_head).json()
            logger.info(search_res)
            potentialInfo = search_res["body"]["list"]
            logger.info(len(potentialInfo))
            self.procDefKey = []
            self.procInstId = []
            supplierCode = []
            self.busiId = []
            for i in range(len(potentialInfo)):
                supplierCode.append(potentialInfo[i]["supplierCode"])
                self.procDefKey.append(potentialInfo[i]["procDefKey"])
                self.procInstId.append(potentialInfo[i]["procInstId"])
                self.busiId.append(potentialInfo[i]["busiId"])
            for q in range(len(potentialInfo)):
                if self.bill_sn == supplierCode[q]:
                    self.procDefKey = self.procDefKey[q]
                    self.procInstId = self.procInstId[q]
                    self.busiId = self.busiId[q]
                continue
            logger.info(f"获取库存发布编号为{self.bill_sn}的审核id的list列表为：{self.procInstId},审核任务类型为{self.procDefKey}")
            passtask_url = "{}/partnermanage/passTask".format(self.SRM_URL)
            passtask_body = {"comment": comment, "operation": operation, "procDefKey": self.procDefKey,
                             "procInstId": self.procInstId, "supplierCode": self.bill_sn}
            logger.info(passtask_body)
            passtask_res = self.srm_rss.post(url=passtask_url, json=passtask_body, headers=self.json_head).json()
            logger.info(passtask_res)
            suc = passtask_res["suc"]
            if suc == True:
                logger.info("审核完成")
        return suc, self.bill_sn
    def settle_goods_audit(self):
        # 库存更新审核对接我的审批
        suc = None
        i = 0
        while True:
            self.uidCreator = self.settle_goods_auditor_acquire(self.settle_goods_procInstId)
            if self.uidCreator != None:
                suc, self.bill_sn = PartnerAudit(self.srm_rss, self.bill_sn, self.bill_sn, self.uidCreator).mian_parther_audit()
            else:
                i += 1
                if i >= 1:
                    SOO_user_params = {'admin_name': "admin", "admin_pwd": "HQ@uat@666", "pro_pwd": "auth221313",
                                       "pro_user": "zhangbajun", "pwd": '123456789', "user": "yemao"}
                    user_params = {"HQCHIP_SOO": SOO_user_params}
                    write_yaml(account_yaml, user_params)
                    target_rss = SOOLogin("uat-srm.huaqiu.com", "partnermanage").target_login()
                    self.srm_rss = target_rss
                    break
        return suc, self.bill_sn

if __name__ == '__main__':
    # 合作商循环导入库存，结合合作商是否合格、合作商的业务设置范围
    target_rss = SOOLogin("uat-srm.huaqiu.com", "partnermanage").target_login()
    supplier_name =PassPartner(target_rss).pass_partner_list()
    logger.info(supplier_name)
    for i in range(len(supplier_name)):
    #     return_value = PassPartner(target_rss,supplier_name[i]).pass_partner_judge()
    #     if return_value == True:
    #         return_business_value = PassPartner(target_rss, supplier_name[i]).pass_partner_list().pass_partner_business()
    #         if return_business_value == True:
    #             UsersPasswordReset(target_rss, supplier_name[i]).password_reset()
                logger.info(f"此时供应商：{supplier_name[i]}")
                if supplier_name[i] not in ("深圳市汇芯微电子有限公司"):
                    bill_sn = SettleGoods(target_rss, supplier_name=supplier_name[i]).settle_goods_file().settle_search_excel().parther_audit_list("通过","pass")
    #         else:
    #             logger.info(f"供应商：{supplier_name[i]}的业务不符合要求")
    #     else:
    #         logger.info(f"供应商：{supplier_name[i]}不符合要求")
    # SettleGoods(target_rss, MOQ=100, min_step_price="1").excel_file_write()

