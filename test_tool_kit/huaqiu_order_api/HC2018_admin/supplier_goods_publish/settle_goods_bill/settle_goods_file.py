import math
import os
import time

import jsonpath
import openpyxl
import requests
import xlrd
import xlwt
import yaml
from xlwt import Workbook

from huaqiu_order_api.HC2018_admin.login.login import Login
from huaqiu_order_api.common.loguru_logger import logger
from huaqiu_order_api.common.my_data import Data
from huaqiu_order_api.common.my_path import settle_goods_dir, yaml_file, dos_agency_sale_dir, dos_consignment_launch_dir, dos_consignment_reprice_dir, dos_futures_launch_dir



class SettleGoods:
    # 库存发布
    def __init__(self, target_rss, supplier_name=None, goods_name=None, provider_name=None, encap=None,
                 min_pack=None, MOQ=None, min_step_price=None, inland_delivery_day=None, profit_point=None, import_type=None
                 ):
        self.dos_rss = target_rss
        self.headers = {"Content-Type": "application/x-www-form-urlencoded",
                        "User-Agent": "Mozilla/5.0 (compatible; HuaQiuRobot-AutoTest/1.0; +https://www.huaqiu.com)"
                        }
        self.headers_json = {"Content-Type": "application/json; charset=UTF-8",
                             "User-Agent": "Mozilla/5.0 (compatible; HuaQiuRobot-AutoTest/1.0; +https://www.huaqiu.com)"
                             }
        self.auth_token = getattr(Data, 'dos_auth_token')
        self.headers_json["Authorization"] = self.headers["Authorization"] = self.auth_token
        self.supplier_name = supplier_name
        self.goods_name = goods_name
        self.provider_name = provider_name
        self.encap = encap
        self.min_pack = min_pack
        self.MOQ = MOQ
        self.min_step_price_CNY = min_step_price
        self.inland_delivery_day = inland_delivery_day
        self.profit_point = profit_point
        self.import_type = import_type
        with open(yaml_file, 'r', encoding='utf-8') as yamlfile:
            data = yaml.load(yamlfile, Loader=yaml.FullLoader)
        self.HC2018_ADMIN_URL = data['HC2018_ADMIN_URL']

    def excel_file_write(self, file_name, file):
        """更新文件里面型号、品牌、封装、最小包装数量、最小起订量、第一阶梯人民币价格、大陆交期"""
        # 打开指定的Excel文件
        workbook = openpyxl.load_workbook(file)
        workbook.encoding = 'utf-8'
        # 选择活动工作表
        sheet = workbook.active
        # 修改Excel文件里面的型号、品牌、封装
        sheet.cell(row=4, column=2).value = self.goods_name
        sheet.cell(row=4, column=3).value = self.provider_name
        sheet.cell(row=4, column=4).value = self.encap
        # 最小包装
        if self.min_pack != None:
            if file_name == "代售库存发布模板.xlsx" or file_name == "期货发布模板.xlsx":
                sheet.cell(row=4, column=6).value = self.min_pack
                if file_name == "代售库存发布模板.xlsx":
                    sheet.cell(row=4, column=10).value = "5000"
                    sheet.cell(row=4, column=13).value = "5000"
            elif file_name == "寄售发布模板.xlsx":
                sheet.cell(row=4, column=6).value = "深圳华秋东莞仓"
                sheet.cell(row=4, column=7).value = "5000"
                sheet.cell(row=4, column=8).value = self.min_pack
            else:
                pass
        # 起订量
        if self.MOQ != None:
            if file_name == "代售库存发布模板.xlsx" or file_name == "期货发布模板.xlsx":
                sheet.cell(row=4, column=7).value = self.MOQ
            elif file_name == "寄售发布模板.xlsx" or file_name == "期货发布模板.xlsx":
                sheet.cell(row=4, column=9).value = self.MOQ
        if self.min_step_price_CNY != None:
            if file_name == "代售库存发布模板.xlsx":
                sheet.cell(row=4, column=8).value = self.min_step_price_CNY
                min_step_price_USD = math.ceil((float(self.min_step_price_CNY) / 7.1921) * 10000) / 10000
                sheet.cell(row=4, column=12).value = min_step_price_USD
                multiple = 1
                for i in range(1, 3):
                     n = 16 + (i-1) * 3 + (i-1)
                     i = i + 1
                     multiple = multiple * i * 10
                     if self.MOQ != None:
                            step_num = int(self.MOQ) * multiple
                            sheet.cell(row=4, column=n + 1).value = step_num
                            sheet.cell(row=4, column=n + 2).value = float(self.min_step_price_CNY) / multiple
                            sheet.cell(row=4, column=n + 3).value = math.ceil((float(self.min_step_price_CNY) / multiple / 7.1921) * 10000) / 10000
                     if self.profit_point != None:
                        if file_name == "代售库存发布模板.xlsx":
                            profit_point_multiple = float(self.profit_point) / 3
                            sheet.cell(row=4, column=9).value = self.profit_point
                            sheet.cell(row=4, column=n + 4).value = float(profit_point_multiple) * (4-i)
            elif file_name == "期货发布模板.xlsx":
                sheet.cell(row=4, column=8).value = self.min_step_price_CNY
                min_step_price_USD = math.ceil((float(self.min_step_price_CNY) / 7.1921) * 10000) / 10000
                sheet.cell(row=4, column=10).value = min_step_price_USD
                multiple = 1
                for i in range(1, 3):
                     n = 14 + (i-1) * 2 + (i-1)
                     i = i + 1
                     multiple = multiple * i * 10
                     if self.MOQ != None:
                            step_num = int(self.MOQ) * multiple
                            sheet.cell(row=4, column=n).value = step_num
                            sheet.cell(row=4, column=n + 1).value = float(self.min_step_price_CNY) / multiple
                            sheet.cell(row=4, column=n + 2).value = math.ceil((float(self.min_step_price_CNY) / multiple / 7.1921) * 10000) / 10000
            elif file_name == "寄售发布模板.xlsx":
                sheet.cell(row=4, column=10).value = self.min_step_price_CNY
                multiple = 1
                for i in range(1, 3):
                    n = 16 + (i-1) * 2
                    i = i + 1
                    multiple = multiple * i * 10
                    if self.MOQ != None:
                            step_num = int(self.MOQ) * multiple
                            sheet.cell(row=4, column=n).value = step_num
                            sheet.cell(row=4, column=n + 1).value = math.ceil((float(self.min_step_price_CNY) / multiple) * 10000) / 10000
            else:
                pass
        if self.inland_delivery_day != None:
                if file_name == "代售库存发布模板.xlsx" or file_name == "期货发布模板.xlsx":
                    if '-' in self.inland_delivery_day:
                        min_delivery_day = self.inland_delivery_day.split('-')[0]
                        max_delivery_day = self.inland_delivery_day.split('-')[1]
                        hk_delivery_day = str(int(min_delivery_day) + 7) + '-' + str(int(max_delivery_day) + 7)
                    else:
                        hk_delivery_day = str(int(self.inland_delivery_day) + 7)
                    if file_name == "代售库存发布模板.xlsx":
                        sheet.cell(row=4, column=11).value = self.inland_delivery_day
                        sheet.cell(row=4, column=14).value = hk_delivery_day
                    elif file_name == "期货发布模板.xlsx":
                        sheet.cell(row=4, column=9).value = self.inland_delivery_day
                        sheet.cell(row=4, column=11).value = hk_delivery_day
                elif file_name == "寄售发布模板.xlsx":
                    sheet.cell(row=4, column=11).value = self.inland_delivery_day
        # 保存修改后的Excel文件
        workbook.save(file)
        workbook.close()
        return self

    def excel_file_write_xls(self, file_name, file):
        workbook = xlrd.open_workbook(file)
        # 创建一个新的工作簿
        new_workbook = xlwt.Workbook()
        # 复制原有工作表到新工作簿
        for sheet_index in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_index)
            new_sheet = new_workbook.add_sheet(sheet.name)

            for row_index in range(sheet.nrows):
                for col_index in range(sheet.ncols):
                    new_sheet.write(row_index, col_index, sheet.cell_value(row_index, col_index))

        # 往新工作簿的特定工作表追加数据
        new_sheet = new_workbook.get_sheet(0)  # 假设我们要操作的是第一个工作表
        # 假设我们要写入的数据是[['new1', 'new2', 'new3'], ['new4', 'new5', 'new6']]
        data = ['', self.goods_name, self.provider_name, self.encap,
         self.min_pack, self.MOQ, self.min_step_price_CNY, self.profit_point, "5000", self.inland_delivery_day, "5000"]
        if self.inland_delivery_day != None:
              if '-' in self.inland_delivery_day:
                  min_delivery_day = self.inland_delivery_day.split('-')[0]
                  max_delivery_day = self.inland_delivery_day.split('-')[1]
                  hk_delivery_day = str(int(min_delivery_day) + 7) + '-' + str(int(max_delivery_day) + 7)
              else:
                  hk_delivery_day = str(int(self.inland_delivery_day) + 7)
              data.append(hk_delivery_day)
        data.append("24+")
        data.append("测试")
        if self.profit_point != None:
            profit_point_multiple = float(self.profit_point) / 3


        for row_index, row in enumerate([data]):
            for col_index, value in enumerate(row):
                new_sheet.write(sheet.nrows + row_index, col_index, value)



        # 保存新工作簿到文件
        new_workbook.save('your_file_new.xls')

    def write_xls(self, filename, sheet_name, data=None):
        # filename = os.path.abspath(filename)
        assert filename.lower().endswith('.xls'), f'不是.xls文件:{filename}'
        if not os.path.exists(os.path.dirname(filename)):
            os.makedirs(os.path.dirname(filename))
        wb: Workbook = xlwt.Workbook(filename)
        sheet: xlwt.Worksheet = wb.add_sheet(sheet_name)
        for row_x, row in enumerate(data):
            for clo_x, clo in enumerate(row):
                sheet.write(row_x, clo_x, clo)
        wb.save(filename)
        print(f"保存到路径:{filename}")
    def settle_goods_file(self, file):
        """上传库存"""
        file_name = os.path.basename(file)
        settle_goods_file_url = "{}/v1/supplier/SuppPublish/uploadFile/?auth_token={}".format(self.HC2018_ADMIN_URL, self.auth_token)
        file = [('file', (file_name, open(file, 'rb'),'multipart/form-source_data.openx mlformats-officedocument.spreadsheetml.sheet'))]
        settle_goods_file_res = self.dos_rss.post(url=settle_goods_file_url, files=file).json()
        logger.info(settle_goods_file_res)
        self.fileInfo = settle_goods_file_res["data"]
        return self.fileInfo
    def settle_search_excel(self):
        """库存发布"""
        supplier_search_url = "{}/v1/common/BasicService/supList".format(self.HC2018_ADMIN_URL)
        supplier_search_body = {"supplier_name": self.supplier_name, "type": 3}
        supplier_search_res = self.dos_rss.post(url=supplier_search_url, json=supplier_search_body,
                                            headers=self.headers_json).json()
        self.supplier_id = supplier_search_res["data"][0]["supplier_id"]
        #提交导入
        self.order_id = None
        if self.import_type == "1":
            # 寄售
            file = dos_consignment_launch_dir
            file_name = os.path.basename(file)
            # self.excel_file_write(file_name, file)
            self.fileInfo = self.settle_goods_file(file)
            settle_goods_file_url = "{}/v1/supplier/SuppPublish/uploadFile/?auth_token={}".format(self.HC2018_ADMIN_URL, self.auth_token)
            file = [('file', (file_name, open(file, 'rb'),'multipart/form-source_data.openxmlformats-officedocument.spreadsheetml.sheet'))]
            body = {"release_type": 1, "type": self.import_type, "supplier_id": self.supplier_id, "file_url": self.fileInfo["file_url"],
                    "file_name": self.fileInfo["file_name"], "fs_file_name": self.fileInfo["fs_file_name"]}
            settle_goods_file_res = self.dos_rss.post(url=settle_goods_file_url, files=file, data=body).json()
            logger.info(settle_goods_file_res)
            self.order_id = settle_goods_file_res["data"]["order_id"]

        elif self.import_type == "2":
            # 代售
            file = dos_agency_sale_dir
            file_name = os.path.basename(file)
            self.excel_file_write(file_name, file)
            settle_goods_file_url = "{}/v1/supplier/SuppPublish/uploadFile/?auth_token={}".format(self.HC2018_ADMIN_URL, self.auth_token)
            file = [('file', (file_name, open(file, 'rb'),'multipart/form-source_data.openxmlformats-officedocument.spreadsheetml.sheet'))]
            body = {"release_type": 1, "type": self.import_type, "supplier_id": self.supplier_id}
            settle_goods_file_res = self.dos_rss.post(url=settle_goods_file_url, files=file, data=body).json()
            logger.info(settle_goods_file_res)
            self.order_id = settle_goods_file_res["data"]["order_id"]
        elif self.import_type == "3":
            # 期货
            file = dos_futures_launch_dir
            file_name = os.path.basename(file)
            self.excel_file_write(file_name, file)
            settle_goods_file_url = "{}/v1/supplier/SuppPublish/uploadFile/?auth_token={}".format(self.HC2018_ADMIN_URL, self.auth_token)
            file = [('file', (file_name, open(file, 'rb'),'multipart/form-source_data.openxmlformats-officedocument.spreadsheetml.sheet'))]
            body = {"release_type": 1, "type": self.import_type, "supplier_id": self.supplier_id}
            settle_goods_file_res = self.dos_rss.post(url=settle_goods_file_url, files=file, data=body).json()
            logger.info(settle_goods_file_res)
            self.order_id = settle_goods_file_res["data"]["order_id"]

        else:
            pass

        return self

    def obtain_order(self, goods_name=None):
        """获取订单编号"""
        # self.supplier_id = 3065
        # self.order_id = 4824
        obtai_order_url = "{}/v1/supplier/SuppPublish/stockLog".format(self.HC2018_ADMIN_URL)
        obtai_order_body = {"supplier_id": self.supplier_id, "type": self.import_type, "status": -10, "is_audtie": 0, "page": 1, "per_page": 100}
        if goods_name != None:
            obtai_order_body["supplier_id"] = ""
            obtai_order_body["goods_name"] = goods_name
        obtai_order_res = self.dos_rss.post(url=obtai_order_url, json=obtai_order_body, headers=self.headers_json).json()
        total = jsonpath.jsonpath(obtai_order_res, "$..total")[0]
        order_id_new = []
        order_sn_new = []
        status_new = []
        if math.ceil(int(total) / 100) > 1:
            for i in range(math.ceil(int(total) / 100)):
                obtai_order_body["page"] = i+1
                obtai_order_res = self.dos_rss.post(url=obtai_order_url, json=obtai_order_body, headers=self.headers_json).json()
                print(obtai_order_res)
                order_id = jsonpath.jsonpath(obtai_order_res, "$..order_id")
                order_sn = jsonpath.jsonpath(obtai_order_res, "$..order_sn")
                dataInfo = jsonpath.jsonpath(obtai_order_res, "$..data")
                status = jsonpath.jsonpath(dataInfo, "$..status")
                order_id_new = order_id_new + order_id
                order_sn_new = order_sn_new + order_sn
                status_new = status_new + status
        else:
            order_id = jsonpath.jsonpath(obtai_order_res, "$..order_id")
            order_sn = jsonpath.jsonpath(obtai_order_res, "$..order_sn")
            dataInfo = jsonpath.jsonpath(obtai_order_res, "$..data")
            status = jsonpath.jsonpath(dataInfo, "$..status")
            order_id_new = order_id_new + order_id
            order_sn_new = order_sn_new + order_sn
            status_new = status_new + status
        order_id_order_sn_json = dict(zip(order_id_new, zip(order_sn_new, status_new)))

        for key in order_id_order_sn_json:
            if key == int(self.order_id):
                self.obtain_order_sn = order_id_order_sn_json[key][0]
                self.obtain_order_status = order_id_order_sn_json[key][1]
                # print(self.obtain_order_sn)
            else:
                pass
        return self
    def order_submit(self):
        """订单提交"""
        print(self.obtain_order_status)
        if self.obtain_order_status == 0:  # 待确认
            order_submit_url = "{}/v1/supplier/SuppPublish/doConfirm".format(self.HC2018_ADMIN_URL)
            order_submit_body = {"order_id": self.order_id, "stock_type": self.import_type, "is_lower_shelf":0,"supplier_id": self.supplier_id}
            order_submit_order_res = self.dos_rss.post(url=order_submit_url, json=order_submit_body, headers=self.headers_json).json()
            print(f"执行结果：{order_submit_order_res}")

        else:
            pass
        print(self.obtain_order_sn)
        setattr(Data, 'consign_sn', self.obtain_order_sn)
        return self

    def mian_self_file_stockup(self):
        self.settle_search_excel()
        self.obtain_order()
        self.order_submit()



if __name__ == '__main__':
    target_rss = Login().login()
    SettleGoods(target_rss, "测试合作库存优化",  "searchV4.16.22", "searchV4", "06031",
                 100, 100, 1.5, "3-7", 15, "1").mian_self_file_stockup()