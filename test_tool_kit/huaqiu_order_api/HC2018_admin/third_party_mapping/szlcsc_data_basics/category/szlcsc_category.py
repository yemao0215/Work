import math
import re

import jsonpath
import pandas
import pandas as pd
import requests
import yaml
from openpyxl.reader.excel import load_workbook

from huaqiu_order_api.common.loguru_logger import logger
from huaqiu_order_api.common.my_data import Data
from huaqiu_order_api.common.my_path import yaml_file, szlcsc_category_dir


class SzlcscCategory:
    def __init__(self, rss=None):
        with open(yaml_file, 'r', encoding='utf-8') as yamlfile:
            data = yaml.load(yamlfile, Loader=yaml.FullLoader)
        self.HC2018_ADMIN_URL = data['HC2018_ADMIN_URL']
        token = getattr(Data, "dos_auth_token")
        self.headers = {"Content-Type": "application/x-www-form-urlencoded", "Authorization": token}
        self.headers_json = {"Content-Type": "application/json; charset=utf-8", "Authorization": token}
        self.rss =rss

    def read_data(self):
        logger.info("开始读取表格内容")
        data = pandas.read_excel(szlcsc_category_dir)
        self.szlcsc_category_en = data["芯灵类目name"]
        self.category_en = data["DOS类目"]
        self.formal_brand_id = data["芯灵id"]
        self.row_count = len(data)
        self.column_count = data.shape[1]
        return self
    def dos_category_search(self, category_en):
        logger.info(category_en)
        search_url = "{}/v1/goods/DgkCategory/findList".format(self.HC2018_ADMIN_URL)
        search_body = {"cat_name": category_en, "search_type": 1, "is_enabled": "-1", "is_self": "-1", "is_show": "-1", "type": "all", "page": 1, "per_page": 15}
        while True:
            i = 0
            search_res = self.rss.post(url=search_url, json=search_body, headers=self.headers_json).json()
            data = jsonpath.jsonpath(search_res, '$..data')[0]
            total = len(data)
            logger.info(data)
            if total > 0:
                for i in range(total):
                    if "_child" in data[i]:
                        logger.info(111)
                        for k in data[i]["_child"]:
                            if "_child" in k:
                                logger.info(116)
                                v = k["_child"]
                                for w in v:
                                    if w["cat_name"] == category_en:
                                        logger.info(117)
                                        cat_id = k['cat_id']
                                        return cat_id

                            else:
                                if k["cat_name"] == category_en:
                                    logger.info(114)
                                    cat_id = k['cat_id']
                                    return cat_id
                                # else:
                                #     logger.info(118)
                                #     i += 1
                                #     self.dos_category_create(category_en)
                                #     search_url = "{}/v1/goods/DgkCategory/findMyOrDelList".format(self.HC2018_ADMIN_URL)
                                #     search_body = {"cat_name": category_en, "search_type": 1, "is_enabled": "0",
                                #                    "is_self": "-1", "is_show": "-1", "type": "my", "page": 1,
                                #                    "per_page": 15}
                                #     search_res = self.rss.post(url=search_url, json=search_body,
                                #                                headers=self.headers_json).json()
                                #     cat_id = jsonpath.jsonpath(search_res, '$..cat_id')
                                #     logger.info(cat_id)
                                #     if cat_id != None:
                                #         self.dos_category_giveaudit(cat_id)
                                #         self.dos_category_audit(category_en)
                                #     if i >= 1:
                                #         break

                    else:
                         logger.info(112)
                         for k in data:
                             if k["cat_name"] == category_en:
                                 logger.info(113)
                                 cat_id = k['cat_id']
                                 return cat_id
            else:
                logger.info(115)
                i += 1
                self.dos_category_create(category_en)
                search_url = "{}/v1/goods/DgkCategory/findMyOrDelList".format(self.HC2018_ADMIN_URL)
                search_body = {"cat_name": category_en, "search_type": 1, "is_enabled": "0", "is_self": "-1", "is_show": "-1", "type": "my", "page": 1, "per_page": 15}
                search_res = self.rss.post(url=search_url, json=search_body, headers=self.headers_json).json()
                cat_id = jsonpath.jsonpath(search_res, '$..cat_id')
                logger.info(cat_id)
                if cat_id != None:
                    self.dos_category_giveaudit(cat_id)
                    self.dos_category_audit(category_en)
                if i >= 1:
                    break


    def dos_category_create(self, category_en):
        """品牌创建"""
        category_create_url = "{}/v1/goods/DgkCategory/insert".format(self.HC2018_ADMIN_URL)
        category_create_body = {"cat": {"cat_name": category_en, "keywords": "", "cat_letter": "L", "is_show": "1",
                             "finance_cate_type": "4", "is_describe": 0, "is_required": 0, "sort_order": "99"}}
        category_create_res = self.rss.post(url=category_create_url, json=category_create_body, headers=self.headers_json)
        # logger.info(category_create_res)
        if category_create_res.status_code == 200:
            logger.info(f"类目新增：{category_en}新增成功，执行结果：{category_create_res.json()}")
        return self

    def dos_category_giveaudit(self, cat_id):
        category_submit_audit_url = "{}/v1/goods/DgkCategory/submitAudit".format(self.HC2018_ADMIN_URL)
        category_submit_audit_body = {"cat_id": cat_id}
        category_submit_audit_res = self.rss.post(url=category_submit_audit_url, data=category_submit_audit_body, headers=self.headers)
        if category_submit_audit_res.status_code == 200:
            logger.info(f"分类id：{cat_id}提审成功，执行结果：{category_submit_audit_res.json()}")
        return self
    def dos_category_audit(self, category_en):
        """品牌审核"""
        category_audit_search_url = "{}/v1/goods/DgkCategory/findAuditList".format(self.HC2018_ADMIN_URL)
        category_audit_search_body = {"cat_name": category_en, "global_audit_status": 1, "is_self": "-1", "page": 1, "per_page": 100}
        while True:
            i = 0
            search_res = self.rss.post(url=category_audit_search_url, json=category_audit_search_body, headers=self.headers_json).json()
            total = jsonpath.jsonpath(search_res, '$..total')
            cat_id = jsonpath.jsonpath(search_res, '$..cat_id')
            logger.info(search_res)
            if int(total[0]) < 1:
                logger.info("已审核")
                return cat_id
            else:
                k = ""
                if cat_id != []:
                    for k in cat_id:
                        logger.info(k)
                        category_audit_url = "{}/v1/goods/DgkCategory/audit".format(self.HC2018_ADMIN_URL)
                        category_audit_body = {"cat_id": k, "status": 1}
                        brand_audit_res = self.rss.post(url=category_audit_url, json=category_audit_body, headers=self.headers_json)
                        if brand_audit_res.status_code == 200:
                            logger.info(f"分类：{category_en}审核成功，执行结果：{brand_audit_res.json()}")

                    i += 1
                    logger.info(i)
                    if i >= 4:
                        break
                # cat_id = k
                # return cat_id

    def write_data(self,filename, sheetname, row, column1, actual):
        """
        在指定行写入数据
        :param row: 行号
        :param actual: 实际结果
        :param result: 最终结果是否通过Fail/Pass
        :return:
        """
        # 同一个workbook对象，如将多个数据写入不同表单，则只有最后一个表单能写入成功---这是openpyxl的特性，无法避免。
        # 要写入不同的表单，须重新再定义一个workbook对象
        other_wb = load_workbook(filename)
        if sheetname is None:
            other_ws = other_wb.active
        else:
            other_ws = other_wb[sheetname]
        if isinstance(row, int) and (2 <= row <= other_ws.max_row):  # 行号为整数，且行号为第2行以后的数据
            other_ws.cell(row=row, column=column1, value=actual)
            other_wb.save(filename)  # 写入成功后保存文件
            other_wb.close()  # openpyxl读数据时不关闭，写数据时可关闭也可不关闭
        else:  # 如果行号不是整数/小于2/大于最大行号了就会报错
            print("传入的行号有误，  请重新写入")
    def main_szlcsc_category(self):
        self.read_data()
        n = 1
        for i in range(self.row_count):
            n = n + 1
            logger.info(i)
            brand_id = self.dos_category_search(self.category_en[i])
            if brand_id != None:
                self.write_data(szlcsc_category_dir, 'Sheet1', n, self.column_count + 1, brand_id)
        return self

    def match_brand_cn(self, string):
        pattern = r'^\d+(\.\d+)?$'
        flags = 0
        try:
            x = float(string)
            # 判断string 是否为NaN, 是则传空值
            if math.isnan(x) == True:
                string = ""
            match_obj = re.compile(pattern, flags).match(str(string))
            return match_obj

        except ValueError:
            logger.info("无法将字符串转换为浮点数")
            return string


if __name__ == '__main__':
    from huaqiu_order_api.HC2018_admin.login.login import Login
    rss = Login().login()
    SzlcscCategory(rss).main_szlcsc_category()