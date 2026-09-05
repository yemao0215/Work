import math
import re

import jsonpath
import pandas
import pandas as pd
import requests
import yaml
from openpyxl.reader.excel import load_workbook

from huaqiu_order_api.common.loguru_logger import logger
from huaqiu_order_api.common.my_data import Data
from huaqiu_order_api.common.my_path import yaml_file, szlcsc_brand_dir


class SzlcscBrand:
    def __init__(self, rss=None):
        with open(yaml_file, 'r', encoding='utf-8') as yamlfile:
            data = yaml.load(yamlfile, Loader=yaml.FullLoader)
        self.HC2018_ADMIN_URL = data['HC2018_ADMIN_URL']
        token = getattr(Data, "dos_auth_token")
        self.headers = {"Content-Type": "application/x-www-form-urlencoded", "Authorization": token}
        self.headers_json = {"Content-Type": "application/json; charset=utf-8", "Authorization": token}
        self.rss =rss

    def read_data(self):
        logger.info("开始读取表格内容")
        data = pandas.read_excel(szlcsc_brand_dir)
        self.szlcsc_brand_en = data["品牌英文"]
        self.szlcsc_brand_cn = data["品牌中文"]
        self.formal_brand_id = data["DOS_brand_id"]
        self.brand_en = data["英文简称"]
        self.brand_cn = data["中文简称"]
        # self.web_display = data["前端显示文本"]
        self.row_count = len(data)
        self.column_count = data.shape[1]
        return self
    def dos_brand_search(self, brand_name, brand_cn):
        logger.info(brand_name)
        search_url = "{}/v1/goods/DgkBrand/brandList".format(self.HC2018_ADMIN_URL)
        search_body = {"brand_name": brand_name, "brand_type": 1, "is_exact": "0", "page": 1, "per_page": 100}
        while True:
            i = 0
            search_res = self.rss.post(url=search_url, json=search_body, headers=self.headers_json).json()
            total = jsonpath.jsonpath(search_res, '$..total')
            logger.info(search_res)
            if int(total[0]) > 0:
                num = math.ceil(int(total[0])/100)
                brand_list = []
                for i in range(num):
                    i = i + 1
                    search_body["page"] = i
                    search_res = self.rss.post(url=search_url, json=search_body, headers=self.headers_json).json()
                    brand_res_list = jsonpath.jsonpath(search_res, '$..list')[0]
                    brand_list = brand_list + brand_res_list
                # logger.info(brand_list)
                for k in brand_list:
                    # logger.info(k['brand_name'])
                    if k['brand_name'].strip() == brand_name.strip():
                        logger.info(111)
                        brand_id = k['brand_id']
                        logger.info(brand_id)
                        return brand_id
                    # elif brand_name in k['brand_name'].strip():
                    #     logger.info(112)
                    #     brand_id = k['brand_id']
                    #     logger.info(brand_id)
                    #     return brand_id
                    # elif brand_name.strip().upper() in k['brand_name'].strip():
                    #     logger.info(113)
                    #     brand_id = k['brand_id']
                    #     logger.info(brand_id)
                    #     return brand_id
                    elif brand_name.strip().upper() == k['brand_name'].strip():
                        logger.info(113)
                        brand_id = k['brand_id']
                        logger.info(brand_id)
                        return brand_id
                    elif brand_name.strip().upper() == k['brand_name'].strip().upper():
                        logger.info(113)
                        brand_id = k['brand_id']
                        logger.info(brand_id)
                        return brand_id
                else:
                    logger.info(117)
                    self.dos_brand_create(brand_name, brand_cn)
                    search_url = "{}/v1/goods/DgkBrand/brandList".format(self.HC2018_ADMIN_URL)
                    search_body = {"brand_name": brand_name, "brand_type": 1, "is_exact": "0", "page": 1,
                                       "per_page": 100, "global_audit_status": ["-1"]}
                    search_res = self.rss.post(url=search_url, json=search_body, headers=self.headers_json).json()
                    brand_id = jsonpath.jsonpath(search_res, '$..brand_id')[0]
                    self.dos_brand_giveaudit(brand_id)
                    self.dos_brand_audit(brand_name, brand_id)
                    return brand_id
            else:
                logger.info(114)
                check_url = "{}/v1/goods/DgkBrand/checkBrandName".format(self.HC2018_ADMIN_URL)
                check_body = {"brand_name": brand_name}
                check_res = self.rss.post(url=check_url, json=check_body, headers=self.headers_json).json()
                if check_res['msg'] == "名称可用":
                    logger.info(115)
                    self.dos_brand_create(brand_name, brand_cn)
                    search_url = "{}/v1/goods/DgkBrand/brandList".format(self.HC2018_ADMIN_URL)
                    search_body = {"brand_name": brand_name, "brand_type": 1, "is_exact": "0", "page": 1, "per_page": 100, "global_audit_status": ["-1"]}
                    search_res = self.rss.post(url=search_url, json=search_body, headers=self.headers_json).json()
                    brand_id = jsonpath.jsonpath(search_res, '$..brand_id')[0]
                    self.dos_brand_giveaudit(brand_id)
                    self.dos_brand_audit(brand_name, brand_id)
                i += 1
                if i >= 1:
                    break


    def dos_brand_create(self, brand_name, brand_cn):
        brand_cn = self.match_brand_cn(brand_cn)
        if brand_cn == None:
            brand_cn = ""
        """品牌创建"""
        brand_create_url = "{}/v1/goods/DgkBrand/brandInsert".format(self.HC2018_ADMIN_URL)
        brand_create_body = {"brand": {"brand_name": brand_name, "brand_cn": brand_cn, "brand_attr": "1", "is_show": "1",
                             "is_hot": "0", "is_new": "0", "location_type": "0", "auth_brand_img_name": "授权证明"}}
        brand_create_res = self.rss.post(url=brand_create_url, json=brand_create_body, headers=self.headers_json)
        # logger.info(brand_create_res)
        if brand_create_res.status_code == 200:
            logger.info(f"品牌英文：{brand_name}新增成功，执行结果：{brand_create_res.json()}")
        return self

    def dos_brand_giveaudit(self, brand_id):
        brand_submit_audit_url = "{}/v1/goods/DgkBrand/giveAudit".format(self.HC2018_ADMIN_URL)
        brand_submit_audit_body = {"ids": brand_id}
        brand_submit_audit_res = self.rss.post(url=brand_submit_audit_url, data=brand_submit_audit_body, headers=self.headers)
        if brand_submit_audit_res.status_code == 200:
            logger.info(f"品牌id：{brand_id}提审成功，执行结果：{brand_submit_audit_res.json()}")
        return self
    def dos_brand_audit(self, brand_name, brand_id):
        """品牌审核"""
        brand_audit_search_url = "{}/v1/goods/DgkBrand/brandAuditList".format(self.HC2018_ADMIN_URL)
        brand_audit_search_body = {"brand_name": brand_name, "status": 1, "page": 1, "per_page": 100}
        while True:
            i = 0
            search_res = self.rss.post(url=brand_audit_search_url, json=brand_audit_search_body, headers=self.headers_json).json()
            total = jsonpath.jsonpath(search_res, '$..total')
            audit_id = jsonpath.jsonpath(search_res, '$..id')
            logger.info(search_res)
            if int(total[0]) < 1:
                logger.info("已审核")
                return brand_id
            else:
                if audit_id != []:
                    for k in audit_id:
                        brand_audit_url = "{}/v1/goods/DgkBrand/AuditPass".format(self.HC2018_ADMIN_URL)
                        brand_audit_body = {"ids": k}
                        brand_audit_res = self.rss.post(url=brand_audit_url, json=brand_audit_body, headers=self.headers_json)
                        if brand_audit_res.status_code == 200:
                            logger.info(f"品牌英文：{brand_name}审核成功，执行结果：{brand_audit_res.json()}")
                    i += 1
                    logger.info(i)
                    if i >= 4:
                        break

    def write_data(self,filename, sheetname, row, column1, actual):
        """
        在指定行写入数据
        :param row: 行号
        :param actual: 实际结果
        :param result: 最终结果是否通过Fail/Pass
        :return:
        """
        # 同一个workbook对象，如将多个数据写入不同表单，则只有最后一个表单能写入成功---这是openpyxl的特性，无法避免。
        # 要写入不同的表单，须重新再定义一个workbook对象
        other_wb = load_workbook(filename)
        if sheetname is None:
            other_ws = other_wb.active
        else:
            other_ws = other_wb[sheetname]
        if isinstance(row, int) and (2 <= row <= other_ws.max_row):  # 行号为整数，且行号为第2行以后的数据
            other_ws.cell(row=row, column=column1, value=actual)
            other_wb.save(filename)  # 写入成功后保存文件
            other_wb.close()  # openpyxl读数据时不关闭，写数据时可关闭也可不关闭
        else:  # 如果行号不是整数/小于2/大于最大行号了就会报错
            print("传入的行号有误，  请重新写入")
    def main_szlcsc_brand(self):
        self.read_data()
        n = 1
        for i in range(self.row_count):
            n = n + 1
            logger.info(i)
            brand_id = self.dos_brand_search(self.brand_en[i], self.brand_cn[i])
            if brand_id != None:
                self.write_data(szlcsc_brand_dir, 'Sheet1', n, self.column_count + 1, brand_id)
        return self

    def match_brand_cn(self, string):
        pattern = r'^\d+(\.\d+)?$'
        flags = 0
        try:
            x = float(string)
            # 判断string 是否为NaN, 是则传空值
            if math.isnan(x) == True:
                string = ""
            match_obj = re.compile(pattern, flags).match(str(string))
            return match_obj

        except ValueError:
            logger.info("无法将字符串转换为浮点数")
            return string

if __name__ == '__main__':
    from huaqiu_order_api.HC2018_admin.login.login import Login
    rss = Login().login()
    SzlcscBrand(rss).main_szlcsc_brand()
    # SzlcscBrand(rss).dos_brand_giveaudit(57923).dos_brand_audit("Hoei", 57923)
    # a = "ams OSRAM"
    # b = "AMS OSRAM"
    # if a.strip().upper() == b.strip():
    #     print(a.strip().upper())
